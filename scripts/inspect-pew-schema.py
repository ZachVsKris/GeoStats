#!/usr/bin/env python3
from __future__ import annotations
import csv, io, re, tempfile, zipfile
from pathlib import Path
from urllib.request import Request, urlopen
from openpyxl import load_workbook

URL='https://www.pewresearch.org/wp-content/uploads/sites/20/2025/06/Religious-Composition-2010-2020-dataset.zip'
req=Request(URL,headers={'User-Agent':'GeoStats Pew schema audit'})
raw=urlopen(req,timeout=180).read()
print('download_bytes',len(raw))
with tempfile.TemporaryDirectory() as d:
    root=Path(d)
    z=root/'pew.zip'; z.write_bytes(raw)
    with zipfile.ZipFile(z) as h:
        print('archive_files')
        for n in h.namelist(): print(' -',n)
        h.extractall(root)
    files=[p for p in root.rglob('*') if p.suffix.lower() in {'.xlsx','.xlsm','.csv'} and not p.name.startswith('~$')]
    print('tabular_files',[(p.name,p.stat().st_size) for p in files])
    for p in files:
        print('\nFILE',p.name)
        if p.suffix.lower()=='.csv':
            rows=list(csv.reader(p.open(encoding='utf-8-sig',newline='')))
            for row in rows[:12]: print('ROW',row[:30])
            continue
        wb=load_workbook(p,read_only=True,data_only=True)
        print('SHEETS',wb.sheetnames)
        for ws in wb.worksheets:
            print('\nSHEET',ws.title,ws.max_row,ws.max_column)
            rows=[]
            for i,row in enumerate(ws.iter_rows(values_only=True)):
                vals=list(row)
                rows.append(vals)
                print('ROW',i+1,vals[:35])
                if i>=14: break
            # Print any header-looking row in first 40 rows in full.
            for i,row in enumerate(ws.iter_rows(values_only=True)):
                vals=[str(v or '') for v in row]
                joined=' | '.join(v.lower() for v in vals)
                if 'country' in joined and ('christian' in joined or 'muslim' in joined):
                    print('HEADER_ROW',i+1)
                    for j,v in enumerate(vals): print(j,repr(v))
                    # sample Mongolia row following header area
                    for rr in ws.iter_rows(min_row=i+2,values_only=True):
                        s=[str(v or '') for v in rr]
                        if any('mongolia'==v.strip().lower() for v in s):
                            print('MONGOLIA_ROW')
                            for j,v in enumerate(rr): print(j,repr(v))
                            break
                    break
