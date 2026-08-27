#!/usr/bin/env python3
"""GeoStats v16.2.6 UN World Urbanization Prospects 2025 country importer.

Recovers the 11 country-level Degree of Urbanization concepts from the official
WUP 2025 F01-F04 tables.  The importer accepts either:

* a normalized CSV with columns ISO3, Country, Year, Category,
  PopulationThousands, PopulationSharePct, PopulationGrowthPct, ShareGrowthPct; or
* a directory containing the official F01-F04 XLSX files.

Official-workbook parsing is deliberately fail-closed.  The parser requires an
ISO3 field, a Degree-of-Urbanization category field, and an explicit 2025 value.
It supports both long (Year + Value) and wide (year columns) layouts, but does
not guess when workbook structure is ambiguous.
"""
from __future__ import annotations

import argparse
import csv
import os
import re
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from data_pipeline.base import WarehouseImporter
from data_pipeline.canonical_countries import CANONICAL_COUNTRY_NAMES, canonical_country_name, country_name_to_iso3
from data_pipeline.models import CandidateDefinition, IndicatorRule, SourceObservation
from data_pipeline.supabase import SupabaseWarehouse

SOURCE_ORG = "United Nations Department of Economic and Social Affairs, Population Division"
SOURCE_DATASET = "World Urbanization Prospects 2025"
SOURCE_PAGE = "https://population.un.org/wup/"
DOWNLOAD_PAGE = "https://population.un.org/wup/downloads"
METHOD = "https://population.un.org/wup/assets/Publications/undesa_pd_2025_wup2025_methodological_report.pdf"
YEAR = 2025

FILES = {
    "population": "WUP2025-F01-Degree-of-Urbanization_Pop_by_category.xlsx",
    "share": "WUP2025-F02-Degree-of-Urbanization_percPop_by_category.xlsx",
    "population_growth": "WUP2025-F03-Degree-of-Urbanization_Pop_rate_by_category.xlsx",
    "share_growth": "WUP2025-F04-Degree-of-Urbanization_percPop_rate_by_category.xlsx",
}
BASE_DOWNLOAD = "https://population.un.org/wup/assets/Download/Countries%20and%20Aggregates/"


def norm(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def parse_float(value: object) -> float | None:
    if value in (None, "", "..", "…", "—", "-"):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def classify_category(value: object) -> str | None:
    text = str(value or "").strip().lower()
    if not text:
        return None
    # Check towns before cities because some labels include explanatory text.
    if "town" in text or "semi-dense" in text or "semidense" in text:
        return "towns"
    if "rural" in text:
        return "rural"
    if "cit" in text:
        return "cities"
    return None


def _canonical_iso3(raw_iso3: object, raw_country: object) -> str | None:
    iso3 = str(raw_iso3 or "").strip().upper()
    if iso3 not in CANONICAL_COUNTRY_NAMES:
        iso3 = country_name_to_iso3(str(raw_country or "")) or ""
    return iso3 if iso3 in CANONICAL_COUNTRY_NAMES else None


def _find_header_row(rows: list[list[object]]) -> tuple[int, dict[str, int]]:
    """Return a single-row long/wide header map when one is unambiguous."""
    for idx, row in enumerate(rows[:80]):
        names = [norm(v) for v in row]
        mapping: dict[str, int] = {}
        for i, name in enumerate(names):
            if name in {"iso3", "iso3code", "iso3alphacode", "alpha3code"}:
                mapping.setdefault("iso3", i)
            if any(token in name for token in ("countryorarea", "countryarea", "location")):
                mapping.setdefault("country", i)
            if name in {"year", "time", "timeperiod"}:
                mapping.setdefault("year", i)
            if "degreeofurbanization" in name or name in {"category", "class", "degurba"}:
                mapping.setdefault("category", i)
            if name in {"value", "estimate", "population", "percentage", "rate"}:
                mapping.setdefault("value", i)
        # Wide tables need year-number columns; long tables need a year column.
        has_year_number = any(str(v).strip().isdigit() and 1950 <= int(str(v).strip()) <= 2050 for v in row if v not in (None, ""))
        if "iso3" in mapping and "category" in mapping and ("year" in mapping or has_year_number):
            return idx, mapping
    raise RuntimeError("Could not locate an unambiguous WUP header with ISO3, Degree-of-Urbanization category, and year information.")


def _parse_xlsx(path: Path) -> list[tuple[str, str, str, int, float]]:
    """Parse one official WUP workbook into (iso3,country,category,year,value)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    parsed: list[tuple[str, str, str, int, float]] = []
    matched_sheet = False
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 250000), values_only=True)]
        try:
            header_idx, mapping = _find_header_row(rows)
        except RuntimeError:
            continue
        header = rows[header_idx]
        matched_sheet = True
        year_columns = {
            i: int(str(value).strip())
            for i, value in enumerate(header)
            if str(value or "").strip().isdigit() and 1950 <= int(str(value).strip()) <= 2050
        }
        long_layout = "year" in mapping
        if long_layout and "value" not in mapping:
            # Some long-format exports name the measure column after the table.
            candidate_cols = [i for i, name in enumerate([norm(v) for v in header]) if any(t in name for t in ("population", "percent", "rate", "value"))]
            candidate_cols = [i for i in candidate_cols if i not in {mapping.get("category"), mapping.get("country"), mapping.get("iso3"), mapping.get("year")}]
            if len(candidate_cols) == 1:
                mapping["value"] = candidate_cols[0]
        if long_layout and "value" not in mapping:
            raise RuntimeError(f"WUP workbook {path.name} has a long-format year column but no unique value column.")
        if not long_layout and not year_columns:
            raise RuntimeError(f"WUP workbook {path.name} has neither a long-format year column nor year-value columns.")
        for row in rows[header_idx + 1 :]:
            if not any(v not in (None, "") for v in row):
                continue
            def cell(key: str) -> object:
                i = mapping.get(key)
                return row[i] if i is not None and i < len(row) else None
            category = classify_category(cell("category"))
            if not category:
                continue
            iso3 = _canonical_iso3(cell("iso3"), cell("country"))
            if not iso3:
                continue
            country = canonical_country_name(iso3, str(cell("country") or iso3))
            if long_layout:
                try:
                    year = int(float(str(cell("year"))))
                except Exception:
                    continue
                value = parse_float(cell("value"))
                if value is not None:
                    parsed.append((iso3, country, category, year, value))
            else:
                for col, year in year_columns.items():
                    if col >= len(row):
                        continue
                    value = parse_float(row[col])
                    if value is not None:
                        parsed.append((iso3, country, category, year, value))
        # An official file should have one usable data sheet.  Do not merge a second
        # matching sheet because notes/annex copies could create duplicate conflicts.
        if parsed:
            break
    if not matched_sheet or not parsed:
        raise RuntimeError(f"No parseable WUP observations found in {path.name}.")
    return parsed


def _load_normalized_csv(path: Path) -> dict[str, dict[tuple[str, int, str], float]]:
    required = {"iso3", "year", "category"}
    metric_columns = {
        "population": "populationthousands",
        "share": "populationsharepct",
        "population_growth": "populationgrowthpct",
        "share_growth": "sharegrowthpct",
    }
    out: dict[str, dict[tuple[str, int, str], float]] = {k: {} for k in metric_columns}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fields = {norm(v): v for v in (reader.fieldnames or [])}
        if not required.issubset(fields):
            raise RuntimeError("Normalized WUP CSV requires ISO3, Year, and Category columns.")
        for metric, normalized_name in metric_columns.items():
            if normalized_name not in fields:
                raise RuntimeError(f"Normalized WUP CSV is missing {normalized_name}.")
        for row in reader:
            mapped = {norm(k): v for k, v in row.items()}
            iso3 = _canonical_iso3(mapped.get("iso3"), mapped.get("country"))
            category = classify_category(mapped.get("category"))
            try:
                year = int(float(str(mapped.get("year"))))
            except Exception:
                continue
            if not iso3 or not category:
                continue
            for metric, normalized_name in metric_columns.items():
                value = parse_float(mapped.get(normalized_name))
                if value is None:
                    continue
                key = (iso3, year, category)
                if key in out[metric] and abs(out[metric][key] - value) > 1e-12:
                    raise RuntimeError(f"Conflicting duplicate WUP observation for {metric} {key}.")
                out[metric][key] = value
    return out


def load_input(input_path: str | None) -> dict[str, dict[tuple[str, int, str], float]]:
    if not input_path:
        raise RuntimeError(
            "WUP 2025 country importer requires --input pointing to either a normalized CSV or a directory containing the official F01-F04 XLSX files; automatic remote bulk download is intentionally not hidden inside validation runs."
        )
    path = Path(input_path)
    if path.is_file() and path.suffix.lower() == ".csv":
        return _load_normalized_csv(path)
    if not path.is_dir():
        raise RuntimeError("--input must be a normalized CSV or a directory containing official WUP 2025 workbooks.")
    data: dict[str, dict[tuple[str, int, str], float]] = {k: {} for k in FILES}
    for metric, filename in FILES.items():
        file_path = path / filename
        if not file_path.exists():
            raise RuntimeError(f"Missing official WUP file: {filename}")
        for iso3, _country, category, year, value in _parse_xlsx(file_path):
            key = (iso3, year, category)
            if key in data[metric] and abs(data[metric][key] - value) > 1e-12:
                raise RuntimeError(f"Conflicting duplicate WUP observation for {metric} {key}.")
            data[metric][key] = value
    return data


SPECS = {
    "city-population": ("Most people living in cities", "Population living in places classified as cities under the harmonized Degree of Urbanization framework.", "thousand people", "total", "high", "population", "cities"),
    "town-population": ("Most people living in towns", "Population living in towns and semi-dense areas under the harmonized Degree of Urbanization framework.", "thousand people", "total", "high", "population", "towns"),
    "rural-population": ("Largest rural population", "Population living in rural areas under the harmonized Degree of Urbanization framework.", "thousand people", "total", "high", "population", "rural"),
    "city-share": ("Highest city population share", "Share of the population living in places classified as cities.", "% of population", "percentage", "high", "share", "cities"),
    "town-share": ("Highest town population share", "Share of the population living in towns and semi-dense areas.", "% of population", "percentage", "high", "share", "towns"),
    "city-growth": ("Fastest city population growth", "Annual rate of change of the population living in cities.", "% per year", "rate", "high", "population_growth", "cities"),
    "town-growth": ("Fastest town population growth", "Annual rate of change of the population living in towns and semi-dense areas.", "% per year", "rate", "high", "population_growth", "towns"),
    "rural-growth": ("Fastest rural population growth", "Annual rate of change of the population living in rural areas.", "% per year", "rate", "high", "population_growth", "rural"),
    "city-share-growth": ("Fastest increase in city population share", "Annual rate of change of the percentage of the population living in cities.", "% per year", "rate", "high", "share_growth", "cities"),
    "town-share-growth": ("Fastest increase in town population share", "Annual rate of change of the percentage of the population living in towns and semi-dense areas.", "% per year", "rate", "high", "share_growth", "towns"),
    "rural-share-decline": ("Fastest decline in rural population share", "Annual rate of change of the percentage of the population living in rural areas; the most negative rate ranks first.", "% per year", "rate", "low", "share_growth", "rural"),
}

FILE_CODE = {"population": "F01", "share": "F02", "population_growth": "F03", "share_growth": "F04"}


class Importer(WarehouseImporter):
    source_organization = SOURCE_ORG
    source_dataset = SOURCE_DATASET
    source_slug = "unwup2025"

    def __init__(self, warehouse, input_path: str | None, dry_run: bool = False):
        super().__init__(warehouse, dry_run=dry_run)
        self.data = load_input(input_path)

    def discover(self) -> list[CandidateDefinition]:
        out = []
        for key, (title, desc, unit, value_type, direction, metric, category) in SPECS.items():
            code = FILE_CODE[metric]
            rule = IndicatorRule(
                key=key,
                title=title,
                description=desc,
                plain_language_description=desc,
                technical_definition=f"{desc} WUP 2025 Degree of Urbanization estimate/rate for {YEAR}, using official table {code}.",
                unit_explanation=unit,
                family="Urbanization",
                icon="🏙️",
                unit=unit,
                value_type=value_type,
                ranking_direction=direction,
                include=(key,),
                min_coverage=170,
                evidence_tier="A",
                source_priority=5,
                specificity_score=98,
                recognizability_score=95,
                understandability_score=94,
                fun_score=86,
            )
            filename = FILES[metric]
            out.append(
                CandidateDefinition(
                    rule,
                    f"WUP2025-{code}:{category}:{YEAR}",
                    title,
                    SOURCE_PAGE,
                    {
                        "source_page_url": SOURCE_PAGE,
                        "methodology_url": METHOD,
                        "download_url": BASE_DOWNLOAD + filename,
                        "dataset_release": "World Urbanization Prospects 2025 Revision",
                        "source_query": {"table": code, "degree_of_urbanization": category, "year": YEAR},
                        "measurementType": "percentage" if value_type == "percentage" else ("total" if value_type == "total" else "rate"),
                        "broadDomain": "human-geography",
                        "knowledgeCluster": "urbanization",
                        "strategyFamily": f"urbanization-{metric}",
                        "manual_review_required": True,
                        "comparison_year_policy": "Use the 2025 estimate/reference year only; do not mix projections into the current comparison snapshot.",
                        "derivation_method": "Direct official WUP table value; no manual filling.",
                        "v16_2_6_content_reviewed": True,
                    },
                )
            )
        return out

    def fetch_observations(self, candidate: CandidateDefinition) -> list[SourceObservation]:
        key = candidate.rule.key
        _title, _desc, _unit, _vt, _direction, metric, category = SPECS[key]
        values = self.data[metric]
        out = []
        for iso3 in CANONICAL_COUNTRY_NAMES:
            value = values.get((iso3, YEAR, category))
            if value is None:
                continue
            if metric == "share" and not (0 <= value <= 100):
                raise RuntimeError(f"WUP share outside 0-100 for {iso3}: {value}")
            out.append(
                SourceObservation(
                    iso3,
                    canonical_country_name(iso3, iso3),
                    YEAR,
                    value,
                    SOURCE_PAGE,
                    f"{candidate.source_indicator_code}:{iso3}",
                    "official",
                    {"table": FILE_CODE[metric], "degree_of_urbanization": category, "reference_year": YEAR},
                )
            )
        return out

    def category_id(self, candidate: CandidateDefinition) -> str:
        return f"unwup2025:{candidate.rule.key}"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Normalized CSV or directory containing official WUP 2025 F01-F04 XLSX files")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SECRET_KEY") or os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not args.dry_run and (not url or not key):
        raise SystemExit("Set SUPABASE_URL and SUPABASE_SECRET_KEY/SUPABASE_SERVICE_ROLE_KEY.")
    warehouse = None if args.dry_run else SupabaseWarehouse(url, key)
    print(Importer(warehouse, args.input, args.dry_run).run())


if __name__ == "__main__":
    main()
