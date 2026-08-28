from __future__ import annotations

import csv
import io
import re
import zipfile
import hashlib
from pathlib import Path
from typing import Iterable


def norm(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _delimited_rows(raw: bytes, *, delimiter: str | None = None) -> list[dict[str, object]]:
    text = raw.decode("utf-8-sig", "replace")
    if delimiter is None:
        sample = text[:20000]
        delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
    return [dict(row) for row in csv.DictReader(io.StringIO(text), delimiter=delimiter)]


def _xlsx_rows(path: Path) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - CI installs importer requirements
        raise RuntimeError("openpyxl is required for official XLSX bulk inputs") from exc
    wb = load_workbook(io.BytesIO(path.read_bytes()), read_only=True, data_only=True)
    out: list[dict[str, object]] = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        buffered: list[tuple[object, ...]] = []
        for row in rows:
            buffered.append(tuple(row))
            if len(buffered) > 80:
                break
        # Find a plausible header row rather than assuming formatted official workbooks start at row 1.
        header_index = None
        for idx, row in enumerate(buffered):
            keys = {norm(v) for v in row if v not in (None, "")}
            if ("year" in keys or any(re.fullmatch(r"20\d{2}", key) for key in keys)) and any(
                key in keys for key in {"country", "country name", "country or area", "location", "iso3", "code"}
            ):
                header_index = idx
                break
        if header_index is None:
            continue
        headers = [str(v or "").strip() for v in buffered[header_index]]
        data_rows = buffered[header_index + 1 :]
        data_rows.extend(tuple(row) for row in rows)
        for row in data_rows:
            values = list(row) + [None] * max(0, len(headers) - len(row))
            record = dict(zip(headers, values[: len(headers)]))
            if any(v not in (None, "") for v in record.values()):
                out.append(record)
    return out


def read_official_rows(path_value: str) -> list[dict[str, object]]:
    """Read an official release file without silently converting/scraping webpages.

    Supported inputs are CSV/TSV/TXT, XLSX and ZIP archives containing CSV/TSV/TXT.
    The release workflow downloads the exact official file first, then passes the local path here.
    """
    path = Path(path_value)
    if not path.exists():
        raise RuntimeError(f"Official bulk input does not exist: {path}")
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _xlsx_rows(path)
    raw = path.read_bytes()
    if suffix == ".zip" or raw[:4] == b"PK\x03\x04":
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            names = set(zf.namelist())
            # XLSX is itself a ZIP. Release workflows intentionally use neutral .bulk
            # filenames, so detect workbook structure rather than trusting extension.
            if "xl/workbook.xml" in names:
                return _xlsx_rows(path)
            members = [name for name in names if Path(name).suffix.lower() in {".csv", ".tsv", ".txt"}]
            if not members:
                raise RuntimeError("Official ZIP contains no CSV/TSV/TXT data member and is not an XLSX workbook.")
            rows: list[dict[str, object]] = []
            for member in sorted(members):
                rows.extend(_delimited_rows(zf.read(member), delimiter="\t" if member.lower().endswith(".tsv") else None))
            return rows
    return _delimited_rows(raw, delimiter="\t" if suffix == ".tsv" else None)


def normalized(row: dict[str, object]) -> dict[str, object]:
    return {norm(key): value for key, value in row.items()}


def first_value(row: dict[str, object], *aliases: str) -> object | None:
    data = normalized(row)
    for alias in aliases:
        value = data.get(norm(alias))
        if value not in (None, ""):
            return value
    return None


def number(value: object) -> float | None:
    if value in (None, ""):
        return None
    text = str(value).strip().replace(",", "")
    if text in {"..", "...", "-", "—", "na", "n/a"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def source_file_sha256(path_value: str) -> str:
    """SHA-256 of the exact official bulk file supplied to an importer."""
    path = Path(path_value)
    if not path.exists():
        return ""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
