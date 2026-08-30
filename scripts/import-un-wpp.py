#!/usr/bin/env python3
"""GeoStats v16.2.6 curated UN World Population Prospects 2024 importer.

Uses the final 2023 estimate year only (never projections). The importer accepts the
UN compact CSV/XLSX download and derives a small set of highly understandable
country comparisons. Every derived statistic uses the same WPP release and year.
"""
from __future__ import annotations
import argparse, csv, io, os, re
from pathlib import Path
from openpyxl import load_workbook
from data_pipeline.base import WarehouseImporter
from data_pipeline.canonical_countries import CANONICAL_COUNTRY_NAMES, canonical_country_name
from data_pipeline.models import CandidateDefinition, IndicatorRule, SourceObservation
from data_pipeline.supabase import SupabaseWarehouse

SOURCE_ORG='United Nations Population Division'
SOURCE_DATASET='World Population Prospects 2024'
SOURCE_PAGE='https://population.un.org/wpp/'
METHOD='https://population.un.org/wpp/Publications/Files/WPP2024_Methodology-Report_Final.pdf'
DOWNLOAD='https://population.un.org/wpp/assets/Excel%20Files/1_Indicator%20(Standard)/EXCEL_FILES/1_General/WPP2024_GEN_F01_DEMOGRAPHIC_INDICATORS_COMPACT.xlsx'
YEAR=2023

def norm(s): return re.sub(r'[^a-z0-9]+','',str(s or '').lower())

def _xlsx_rows(raw: bytes):
    """Yield rows from the WPP compact workbook without relying on a fixed header row."""
    wb=load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet_names=[]
    if 'Estimates' in wb.sheetnames:
        sheet_names.append('Estimates')
    sheet_names.extend(name for name in wb.sheetnames if name not in sheet_names)
    try:
        for sheet_name in sheet_names:
            ws=wb[sheet_name]
            rows=ws.iter_rows(values_only=True)
            headers=None
            for row in rows:
                vals=['' if value is None else value for value in row]
                normalized={norm(value) for value in vals if value not in (None,'')}
                has_iso=bool(normalized & {'iso3alphacode','iso3code','iso3'})
                if has_iso and 'year' in normalized:
                    headers=[str(value).strip() if value not in (None,'') else '' for value in vals]
                    break
            if headers is None:
                continue
            for row in rows:
                vals=['' if value is None else value for value in row]
                if not any(value not in (None,'') for value in vals):
                    continue
                if len(vals)<len(headers): vals=vals+['']*(len(headers)-len(vals))
                yield dict(zip(headers, vals[:len(headers)]))
            return
    finally:
        wb.close()
    raise RuntimeError('Could not find the WPP compact demographic-indicator sheet/header.')

def load(path_or_url: str):
    from data_pipeline.http import HttpClient
    raw=Path(path_or_url).read_bytes() if Path(path_or_url).exists() else HttpClient(timeout=240,retries=5,user_agent='GeoStats/16.2.6 WPP').get_bytes(path_or_url)
    rows=list(_xlsx_rows(raw)) if raw[:2]==b'PK' else list(csv.DictReader(io.StringIO(raw.decode('utf-8-sig','replace'))))
    out=[]
    for row in rows:
        m={norm(k):v for k,v in row.items()}
        try: year=int(float(str(m.get('year'))))
        except: continue
        if year!=YEAR: continue
        iso3=str(m.get('iso3alphacode') or m.get('iso3code') or m.get('iso3') or '').upper().strip()
        if len(iso3)!=3 or iso3 not in CANONICAL_COUNTRY_NAMES: continue
        def val(*names):
            for n in names:
                x=m.get(norm(n))
                if x not in (None,''):
                    try:return float(str(x).replace(',',''))
                    except:pass
            return None
        total=val('TPopulation1July','TotalPopulation1July','TPopulation1JulyThousands','Total Population, as of 1 July (thousands)')
        male=val('PopMale1July','MalePopulation1July','Male Population, as of 1 July (thousands)')
        female=val('PopFemale1July','FemalePopulation1July','Female Population, as of 1 July (thousands)')
        metrics={
          'male-share': (100*male/total if male is not None and total and total>0 else None),
          'female-share': (100*female/total if female is not None and total and total>0 else None),
          'sex-ratio': val('SexRatio','Population Sex Ratio, as of 1 July (males per 100 females)'),
          'median-age': val('MedianAgePop','MedianAge','Median Age, as of 1 July (years)'),
          'population-density': val('PopDensity','PopulationDensity','Population Density, as of 1 July (persons per square km)'),
          'population-growth': val('PopGrowthRate','PopulationGrowthRate','Population Growth Rate (percentage)'),
          'fertility': val('TFR','TotalFertilityRate','Total Fertility Rate (live births per woman)'),
          'life-expectancy': val('LEx','LifeExpectancy','Life Expectancy at Birth, both sexes (years)'),
          'female-life-expectancy': val('LExFemale','FemaleLifeExpectancy','Female Life Expectancy at Birth (years)'),
          'male-life-expectancy': val('LExMale','MaleLifeExpectancy','Male Life Expectancy at Birth (years)'),
          'infant-mortality': val('IMR','InfantMortalityRate','Infant Mortality Rate (infant deaths per 1,000 live births)'),
          'natural-change-rate': val('NatChangeRT','RateNaturalChange','RateofNaturalChange','CrudeRateNaturalChange','Rate of Natural Change (per 1,000 population)'),
          'birth-rate': val('CBR','CrudeBirthRate','Crude Birth Rate (births per 1,000 population)'),
          'death-rate': val('CDR','CrudeDeathRate','Crude Death Rate (deaths per 1,000 population)'),
          'net-migration-rate': val('CNMR','NetMigrationRate','CrudeNetMigrationRate','Net Migration Rate (per 1,000 population)'),
          'sex-ratio-at-birth': val('SRB','SexRatioatBirth','SexRatioAtBirthMalesPer100FemaleBirths','Sex Ratio at Birth (males per 100 female births)'),
          'mean-age-childbearing': val('MACB','MeanAgeChildbearing','MeanAgeofChildbearing','Mean Age Childbearing (years)'),
        }
        if metrics['female-life-expectancy'] is not None and metrics['male-life-expectancy'] is not None:
            metrics['female-life-expectancy-advantage']=metrics['female-life-expectancy']-metrics['male-life-expectancy']
        out.append((iso3,metrics))
    return out

SPECS={
 'highest-male-share':('Highest male share of population','Share of the population that is male.','% of population','percentage','high','Population'),
 'highest-female-share':('Highest female share of population','Share of the population that is female.','% of population','percentage','high','Population'),
 'highest-sex-ratio':('Most men per 100 women','Male population per 100 female population.','men per 100 women','rate','high','Population'),
 'lowest-sex-ratio':('Most women relative to men','Lowest number of men per 100 women.','men per 100 women','rate','low','Population'),
 'highest-median-age':('Highest median age','Median age of the population.','years','other','high','Population'),
 'lowest-median-age':('Lowest median age','Median age of the population.','years','other','low','Population'),
 'lowest-pop-density':('Lowest population density','Population per square kilometer.','people/km²','rate','low','Population'),
 'fastest-pop-decline':('Fastest population decline','Annual population growth rate; the most negative rate ranks first.','% per year','rate','low','Population'),
 'lowest-fertility':('Lowest fertility rate','Average number of births per woman under current age-specific fertility rates.','births per woman','rate','low','Population'),
 'highest-life-expectancy':('Highest life expectancy','Life expectancy at birth for both sexes.','years','other','high','Health'),
 'female-life-expectancy-advantage':('Largest female life-expectancy advantage','Female life expectancy minus male life expectancy at birth.','years','other','high','Population'),
 'highest-natural-change-rate':('Fastest natural population increase','Crude rate of natural population increase: births minus deaths, per 1,000 people.','per 1,000 people','rate','high','Population'),
 'highest-birth-rate':('Highest birth rate','Crude birth rate per 1,000 people.','births per 1,000 people','rate','high','Population'),
 'lowest-death-rate':('Fewest annual deaths per 1,000 people','Deaths during the year for every 1,000 people in the population; this is an annual rate, not a daily count.','deaths per 1,000 people','rate','low','Population'),
 'highest-net-migration-rate':('Highest net migration rate','Net migration rate per 1,000 people.','per 1,000 people','rate','high','Population'),
 'highest-sex-ratio-at-birth':('Most boys born per 100 girls','Male births per 100 female births.','boys per 100 girls','rate','high','Population'),
 'highest-mean-age-childbearing':('Oldest average age of mothers at childbirth','Mean age of childbearing.','years','other','high','Population'),
 'highest-male-life-expectancy':('Highest male life expectancy','Male life expectancy at birth.','years','other','high','Health'),
 'highest-female-life-expectancy':('Highest female life expectancy','Female life expectancy at birth.','years','other','high','Health'),
}
KEY_METRIC={
 'highest-male-share':'male-share','highest-female-share':'female-share','highest-sex-ratio':'sex-ratio','lowest-sex-ratio':'sex-ratio',
 'highest-median-age':'median-age','lowest-median-age':'median-age','highest-pop-density':'population-density','lowest-pop-density':'population-density',
 'fastest-pop-decline':'population-growth','lowest-fertility':'fertility','highest-life-expectancy':'life-expectancy',
 'lowest-infant-mortality':'infant-mortality','female-life-expectancy-advantage':'female-life-expectancy-advantage',
 'highest-natural-change-rate':'natural-change-rate','highest-birth-rate':'birth-rate','lowest-death-rate':'death-rate',
 'highest-net-migration-rate':'net-migration-rate','highest-sex-ratio-at-birth':'sex-ratio-at-birth',
 'highest-mean-age-childbearing':'mean-age-childbearing','highest-male-life-expectancy':'male-life-expectancy',
 'highest-female-life-expectancy':'female-life-expectancy'
}
OFFICIAL_SERIES_NAME={
 'male-share':'Male population share (derived from WPP male and total population)',
 'female-share':'Female population share (derived from WPP female and total population)',
 'sex-ratio':'men per 100 women',
 'median-age':'Median Age, as of 1 July (years)',
 'population-density':'Population Density, as of 1 July (persons per square km)',
 'population-growth':'Annual growth rate (%)',
 'fertility':'Total Fertility Rate (live births per woman)',
 'life-expectancy':'Life Expectancy at Birth, both sexes (years)',
 'female-life-expectancy':'Female Life Expectancy at Birth (years)',
 'male-life-expectancy':'Male Life Expectancy at Birth (years)',
 'infant-mortality':'Infant Mortality Rate (infant deaths per 1,000 live births)',
 'natural-change-rate':'Rate of Natural Change (per 1,000 population)',
 'birth-rate':'Crude Birth Rate (births per 1,000 population)',
 'death-rate':'Crude Death Rate (deaths per 1,000 population)',
 'net-migration-rate':'Net Migration Rate (per 1,000 population)',
 'sex-ratio-at-birth':'Sex Ratio at Birth (males per 100 female births)',
 'mean-age-childbearing':'Mean Age Childbearing (years)',
 'female-life-expectancy-advantage':'Female minus male life expectancy at birth (derived from WPP sex-specific series)',
}
OFFICIAL_COLUMN_NAME={
 'sex-ratio':'Population Sex Ratio, as of 1 July (males per 100 females)',
 'population-growth':'Population Growth Rate (percentage)',
}
class Importer(WarehouseImporter):
 source_organization=SOURCE_ORG; source_dataset=SOURCE_DATASET; source_slug='unwpp'
 def __init__(self,warehouse,input_path=DOWNLOAD,dry_run=False): super().__init__(warehouse,dry_run=dry_run); self.rows=load(input_path)
 def discover(self):
  out=[]
  for key,(title,desc,unit,vtype,direction,family) in SPECS.items():
   rule=IndicatorRule(key=key,title=title,description=desc,plain_language_description=desc,technical_definition=f'{desc} WPP 2024 estimate for {YEAR}.',unit_explanation=unit,family=family,icon='👥',unit=unit,value_type=vtype,ranking_direction=direction,include=(key,),min_coverage=180,evidence_tier='A',source_priority=4,specificity_score=98,recognizability_score=97,understandability_score=98,fun_score=95)
   metric=KEY_METRIC[key]
   metadata={'source_page_url':SOURCE_PAGE,'download_url':DOWNLOAD,'methodology_url':METHOD,'dataset_release':'World Population Prospects 2024','source_query':{'year':YEAR,'metric':metric,'variant':'estimate'},'minimum_year':YEAR,'measurementType':('share' if vtype=='percentage' else 'per_capita' if vtype=='per_capita' else 'total' if vtype=='total' else 'other'),'broadDomain':'population','knowledgeCluster':'demographics','strategyFamily':key,'v16_2_6_content_reviewed':True,'license_name':'CC BY 3.0 IGO'}
   if metric in OFFICIAL_COLUMN_NAME: metadata['officialSourceColumn']=OFFICIAL_COLUMN_NAME[metric]
   out.append(CandidateDefinition(rule,f'WPP2024:{metric}:{YEAR}',OFFICIAL_SERIES_NAME[metric],SOURCE_PAGE,metadata))
  return out
 def fetch_observations(self,c):
  metric=KEY_METRIC[c.rule.key]; out=[]
  for iso3,vals in self.rows:
   v=vals.get(metric)
   if v is not None: out.append(SourceObservation(iso3,canonical_country_name(iso3,iso3),YEAR,float(v),SOURCE_PAGE,f'{metric}:{iso3}:{YEAR}','estimated'))
  return out
 def category_id(self,c): return f'unwpp:{c.rule.key}'

def main():
 ap=argparse.ArgumentParser(); ap.add_argument('--input',default=DOWNLOAD); ap.add_argument('--dry-run',action='store_true'); a=ap.parse_args(); u=os.getenv('SUPABASE_URL'); k=os.getenv('SUPABASE_SECRET_KEY') or os.getenv('SUPABASE_SERVICE_ROLE_KEY')
 if not a.dry_run and(not u or not k): raise SystemExit('Set Supabase secrets.')
 print(Importer(None if a.dry_run else SupabaseWarehouse(u,k),a.input,a.dry_run).run())
if __name__=='__main__': main()
