#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import os
import re
import tempfile
import zipfile
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from data_pipeline.base import WarehouseImporter
from data_pipeline.canonical_countries import canonical_country_name, country_name_to_iso3
from data_pipeline.models import CandidateDefinition, IndicatorRule, SourceObservation
from data_pipeline.supabase import SupabaseWarehouse

SOURCE_URL = "https://www.pewresearch.org/wp-content/uploads/sites/20/2025/06/Religious-Composition-2010-2020-dataset.zip"
SOURCE_PAGE = "https://www.pewresearch.org/dataset/dataset-of-global-religious-composition-estimates-for-2010-and-2020/"
FEATURE_PAGE = "https://www.pewresearch.org/religion/feature/religious-composition-by-country-2010-2020/"
METHODOLOGY_URL = "https://www.pewresearch.org/religion/2025/06/09/how-the-global-religious-landscape-changed-from-2010-to-2020-methodology/"
DATASET_RELEASE = "Pew global religious composition estimates for 2010 and 2020"
REFERENCE_YEAR = 2020

OTHER_RELIGIONS_DEFINITION = (
    "Pew’s ‘other religions’ category includes Baha’is, Daoists/Taoists, Jains, "
    "Shintoists, Sikhs, Wiccans, Zoroastrians and many smaller groups, including "
    "some folk or traditional religions."
)
OTHER_RELIGIONS_TITLE = "Highest % following religions outside the five major groups"

GROUPS = {
    "christian": ("Christian", "✝️", ("christian",)),
    "muslim": ("Muslim", "☪️", ("muslim",)),
    "hindu": ("Hindu", "🕉️", ("hindu",)),
    "buddhist": ("Buddhist", "☸️", ("buddhist",)),
    "jewish": ("Jewish", "✡️", ("jewish", "jews")),
    "other-religions": ("other religions", "🛕", ("other religions", "all other religions", "other religion")),
    "unaffiliated": ("religiously unaffiliated", "◯", ("unaffiliated", "no religion")),
}


def _title_label(label: str) -> str:
    return "other-religions" if label == "other-religions" else label


def group_rule(key: str, label: str, icon: str, measure: str) -> IndicatorRule:
    if measure == "share":
        if key == "other-religions":
            title = OTHER_RELIGIONS_TITLE
        elif key == "unaffiliated":
            title = "Highest % of population with no religious affiliation"
        else:
            title = f"Highest % of population that is {_title_label(label)}"
        description = f"Estimated percentage of the population identifying as {label} in 2020."
        unit = "% of population"
        value_type = "percentage"
    else:
        title = "Largest population following other religions" if key == "other-religions" else f"Largest {_title_label(label)} population"
        description = f"Estimated number of people identifying as {label} in 2020."
        unit = "people"
        value_type = "total"
    if key == "other-religions":
        description = (
            "Religions other than Christianity, Islam, Hinduism, Buddhism or Judaism."
            if measure == "share" else f"{description} {OTHER_RELIGIONS_DEFINITION}"
        )
    technical_definition = (
        "Pew Research Center estimate synthesized from censuses, surveys, population registers "
        "and demographic estimation. Results are estimates, not exact counts."
    )
    if key == "other-religions":
        technical_definition = f"{technical_definition} {OTHER_RELIGIONS_DEFINITION}"
    return IndicatorRule(
        key=f"{key}-{measure}", title=title, description=description,
        plain_language_description=description,
        technical_definition=technical_definition,
        unit_explanation=unit, family="Religion", icon=icon, unit=unit,
        value_type=value_type, ranking_direction="high", include=(key,), min_coverage=160,
        evidence_tier="B", source_priority=18, specificity_score=97,
        recognizability_score=97, understandability_score=97, fun_score=95,
        objective_status="objective", modeled_hint=1.0,
    )


def diversity_rule() -> IndicatorRule:
    description = "Pew Research Center’s 2020 Religious Diversity Index score."
    return IndicatorRule(
        key="religious-diversity", title="Most religiously diverse", description=description,
        plain_language_description=description,
        technical_definition="Pew Religious Diversity Index based on the distribution of seven religious groups.",
        unit_explanation="Religious Diversity Index", family="Religion", icon="🕊️",
        unit="diversity index", value_type="index", ranking_direction="high",
        include=("diversity",), min_coverage=160, evidence_tier="B", source_priority=18,
        specificity_score=96, recognizability_score=96, understandability_score=96,
        fun_score=92, objective_status="objective", modeled_hint=1.0,
    )


RULES = [
    group_rule(key, label, icon, measure)
    for key, (label, icon, _aliases) in GROUPS.items()
    for measure in ("population", "share")
] + [diversity_rule()]


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9%]+", " ", str(value or "").lower()).strip()


def _download(url: str) -> bytes:
    request = Request(url, headers={"User-Agent": "GeoStats/16.2.7 Pew religion importer"})
    with urlopen(request, timeout=180) as response:
        return response.read()


def _extract_input(path_or_url: str | None) -> Path:
    if path_or_url and Path(path_or_url).exists():
        return Path(path_or_url)
    raw = _download(path_or_url or SOURCE_URL)
    temporary = Path(tempfile.mkdtemp(prefix="geostats-pew-"))
    archive = temporary / "religion.zip"
    archive.write_bytes(raw)
    with zipfile.ZipFile(archive) as handle:
        handle.extractall(temporary)
    candidates = [p for p in temporary.rglob("*") if p.suffix.lower() in {".xlsx", ".xlsm", ".csv"} and not p.name.startswith("~$")]
    if not candidates:
        raise RuntimeError("Pew download did not contain an XLSX or CSV data file.")
    # Prefer the workbook because it contains counts, percentages and diversity
    # in one authoritative release. Falling back to a CSV remains supported.
    workbooks = [p for p in candidates if p.suffix.lower() in {".xlsx", ".xlsm"}]
    return max(workbooks or candidates, key=lambda p: p.stat().st_size)


def _find_header(rows: list[list[Any]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows[:50]):
        headers = [_norm(value) for value in row]
        combined = " | ".join(headers)
        if "country" in combined and ("christian" in combined or "muslim" in combined):
            return index, headers
    raise RuntimeError("Could not locate the Pew country-data header row.")


def _country_column(headers: list[str]) -> int:
    for index, header in enumerate(headers):
        if header in {"country", "country territory", "country or territory", "name"}:
            return index
    for index, header in enumerate(headers):
        if "country" in header and "code" not in header:
            return index
    raise RuntimeError("Could not locate the country column in the Pew dataset.")


def _year_column(headers: list[str]) -> int | None:
    return next((i for i, h in enumerate(headers) if h in {"year", "reference year"}), None)


def _matches_alias(header: str, aliases: tuple[str, ...]) -> bool:
    return any(alias in header for alias in aliases)


def _resolve_group_column(headers: list[str], aliases: tuple[str, ...]) -> int | None:
    exact = [i for i, header in enumerate(headers) if any(header == alias for alias in aliases)]
    if exact:
        return exact[0]
    matches = [i for i, header in enumerate(headers) if _matches_alias(header, aliases)]
    return matches[0] if matches else None


def _resolve_column(headers: list[str], aliases: tuple[str, ...], measure: str) -> int | None:
    """Resolve explicit mixed-format headers used by legacy/supplied files.

    The official 2025 Pew workbook deliberately reuses generic headers such as
    `Christians` on separate count and percentage sheets. Those sheets are
    handled by source_kind in _parse_rows and never pass through this heuristic.
    """
    candidates: list[tuple[int, int]] = []
    for index, header in enumerate(headers):
        if not _matches_alias(header, aliases):
            continue
        score = 10 + (6 if "2020" in header else 0) - (8 if "2010" in header and "2020" not in header else 0)
        share_tokens = ("percent", "percentage", "share", "pct", "%")
        count_tokens = ("number", "count", "population", "people", "thousand")
        if measure == "share":
            score += 8 if any(t in header for t in share_tokens) else -4
            score -= 8 if any(t in header for t in count_tokens) else 0
        else:
            score += 7 if any(t in header for t in count_tokens) else 1
            score -= 12 if any(t in header for t in share_tokens) else 0
        candidates.append((score, index))
    return max(candidates)[1] if candidates else None


def _resolve_diversity(headers: list[str]) -> int | None:
    candidates=[]
    for i,h in enumerate(headers):
        if ("diversity" in h or "rdi" in h) and "rank" not in h and "level" not in h:
            candidates.append((15 + (5 if "index" in h or "score" in h or "rdi" in h else 0) + (4 if "2020" in h else 0), i))
    return max(candidates)[1] if candidates else None


def _as_number(value: Any) -> float | None:
    if value in (None, "", "—", "-", "NA", "N/A"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(",", "").replace("%", "").replace("<", "").replace(">", "")
    try: return float(cleaned)
    except ValueError: return None


def _as_percent(value: Any) -> float | None:
    number = _as_number(value)
    if number is None: return None
    if 0 <= number <= 1: number *= 100
    return number if 0 <= number <= 100 else None


def _religious_diversity_from_shares(item: dict[str, Any]) -> float | None:
    shares=[]
    for key in GROUPS:
        value=item.get(f"{key.replace('-', '_')}_share")
        if value is None: return None
        shares.append(float(value)/100.0)
    concentration=sum(value*value for value in shares)
    return max(0.0,min(10.0,(1.0-concentration)*11.6))


def _source_kind(name: str) -> str:
    normalized = _norm(name)
    if "unrounded count" in normalized or "rounded count" in normalized:
        return "counts"
    if "percentage" in normalized or "percent" in normalized:
        return "percentages"
    if "diversity" in normalized:
        return "diversity"
    return "mixed"


def _parse_rows(rows: list[list[Any]], *, source_kind: str = "mixed") -> list[dict[str, Any]]:
    header_index, headers = _find_header(rows)
    country_index, year_index = _country_column(headers), _year_column(headers)
    columns: dict[tuple[str,str], int | None] = {}
    for key, (_label, _icon, aliases) in GROUPS.items():
        if source_kind == "counts":
            columns[(key,"population")] = _resolve_group_column(headers, aliases)
            columns[(key,"share")] = None
        elif source_kind == "percentages":
            columns[(key,"population")] = None
            columns[(key,"share")] = _resolve_group_column(headers, aliases)
        elif source_kind == "diversity":
            columns[(key,"population")] = None
            columns[(key,"share")] = None
        else:
            columns[(key,"share")] = _resolve_column(headers, aliases, "share")
            columns[(key,"population")] = _resolve_column(headers, aliases, "population")
    diversity_column = _resolve_diversity(headers) if source_kind in {"mixed","diversity"} else None
    if not any(v is not None for v in columns.values()) and diversity_column is None:
        raise RuntimeError("No 2020 religious composition columns were found.")
    parsed=[]
    for row in rows[header_index+1:]:
        if country_index >= len(row): continue
        country=str(row[country_index] or "").strip()
        if not country: continue
        if year_index is not None and year_index < len(row):
            y=_as_number(row[year_index])
            if y is not None and int(y) != REFERENCE_YEAR: continue
        iso3=country_name_to_iso3(country)
        if not iso3: continue
        item={"iso3":iso3,"country":canonical_country_name(iso3,country)}
        for (key,measure), col in columns.items():
            if col is None or col >= len(row): continue
            if measure == "share":
                value=_as_number(row[col]) if source_kind == "percentages" else _as_percent(row[col])
                if value is not None and not 0 <= value <= 100: value=None
            else:
                value=_as_number(row[col])
                if value is not None and "thousand" in headers[col]: value *= 1000
            item[f"{key.replace('-', '_')}_{measure}"]=value
        if diversity_column is not None and diversity_column < len(row):
            item["religious_diversity"]=_as_number(row[diversity_column])
        if any(v is not None for k,v in item.items() if k not in {"iso3","country"}): parsed.append(item)
    if len(parsed) < 100: raise RuntimeError(f"Only {len(parsed)} GeoStats countries were parsed; expected at least 100.")
    return parsed


def _validate_measure_semantics(rows: list[dict[str, Any]]) -> None:
    """Fail loudly if totals and percentages become conflated again."""
    for key in GROUPS:
        prefix=key.replace('-', '_')
        populations=[float(row[f"{prefix}_population"]) for row in rows if row.get(f"{prefix}_population") is not None]
        shares=[float(row[f"{prefix}_share"]) for row in rows if row.get(f"{prefix}_share") is not None]
        if len(populations) < 100 or len(shares) < 100:
            raise RuntimeError(f"Pew {key} is missing count/share coverage: counts={len(populations)}, shares={len(shares)}")
        if max(populations) < 1_000_000:
            raise RuntimeError(f"Pew {key} population series looks percentage-like; maximum is only {max(populations):g} people.")
        if min(shares) < 0 or max(shares) > 100:
            raise RuntimeError(f"Pew {key} share series is outside 0-100%.")
    # Regression for the production bug: a count column and percentage column
    # must not be the same series merely because Pew reuses column headers.
    for key in GROUPS:
        prefix=key.replace('-', '_')
        comparable=[row for row in rows if row.get(f"{prefix}_population") is not None and row.get(f"{prefix}_share") is not None and float(row[f"{prefix}_population"]) > 100 and float(row[f"{prefix}_share"]) > 0]
        identical=sum(abs(float(row[f"{prefix}_population"])-float(row[f"{prefix}_share"])) < 1e-9 for row in comparable)
        if comparable and identical / len(comparable) > 0.10:
            raise RuntimeError(f"Pew {key} population/share series are suspiciously identical for {identical}/{len(comparable)} countries.")


def _load_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower()==".csv":
        with path.open(newline="",encoding="utf-8-sig") as handle:
            parsed=_parse_rows([list(r) for r in csv.reader(handle)],source_kind=_source_kind(path.name))
        # A single official CSV contains only one measure family, so semantic
        # cross-validation is performed only for the combined workbook.
        return parsed

    workbook=load_workbook(path,read_only=True,data_only=True)
    merged: dict[str, dict[str, Any]] = {}
    errors=[]
    for sheet in workbook.worksheets:
        kind=_source_kind(sheet.title)
        rows=[list(r) for r in sheet.iter_rows(values_only=True)]
        try:
            parsed=_parse_rows(rows,source_kind=kind)
        except RuntimeError as error:
            errors.append(f"{sheet.title}: {error}")
            continue
        for item in parsed:
            iso3=str(item["iso3"])
            target=merged.setdefault(iso3,{"iso3":iso3,"country":item["country"]})
            for key,value in item.items():
                if key not in {"iso3","country"} and value is not None:
                    target[key]=value
    result=sorted(merged.values(),key=lambda item:str(item["iso3"]))
    if len(result)<100:
        raise RuntimeError(f"Only {len(result)} merged GeoStats countries were parsed. "+"; ".join(errors[:5]))
    for item in result:
        if item.get("religious_diversity") is None:
            item["religious_diversity"]=_religious_diversity_from_shares(item)
    _validate_measure_semantics(result)
    return result


class PewReligionImporter(WarehouseImporter):
    source_organization="Pew Research Center"
    source_dataset="Global Religious Composition Estimates for 2010 and 2020"
    source_slug="pewreligion"
    def __init__(self, warehouse: SupabaseWarehouse|None, *, input_path:str|None=None, dry_run:bool=False):
        super().__init__(warehouse,dry_run=dry_run); self.input_path=input_path; self._rows=None
    def rows(self):
        if self._rows is None: self._rows=_load_rows(_extract_input(self.input_path))
        return self._rows
    def discover(self):
        available=set().union(*(row.keys() for row in self.rows()))
        result=[]
        for item in RULES:
            metric=item.key.replace("-","_")
            if metric not in available: continue
            group_key=item.key.rsplit("-",1)[0] if item.key != "religious-diversity" else "diversity"
            metadata={
                "source_page_url":FEATURE_PAGE,"exact_query_url":FEATURE_PAGE,"download_url":SOURCE_URL,
                "dataset_release":DATASET_RELEASE,"minimum_year":2020,"methodology_url":METHODOLOGY_URL,
                "source_query":{"referenceYear":2020,"metric":metric,"estimate":True},
                "broadDomain":"culture","knowledgeCluster":"religious-composition",
                "strategyFamily":f"religion-{group_key}","manual_review_required":True,"estimatedData":True,
                "diversityDerivation":"Pew modified Herfindahl-Hirschman method from seven 2020 shares when official RDI column is absent" if item.key == "religious-diversity" else None,
            }
            if group_key == "other-religions":
                metadata["groupDefinition"] = OTHER_RELIGIONS_DEFINITION
            result.append(CandidateDefinition(item,f"PEW_RELIGION_2020_{metric.upper()}",item.title,FEATURE_PAGE,metadata))
        return result
    def category_id(self,candidate): return f"pew-religion:{candidate.rule.key}"
    def fetch_observations(self,candidate):
        metric=candidate.rule.key.replace("-","_"); out=[]
        for row in self.rows():
            value=row.get(metric)
            if value is None: continue
            out.append(SourceObservation(str(row["iso3"]),str(row["country"]),REFERENCE_YEAR,float(value),FEATURE_PAGE,
                f"{row['iso3']}:{REFERENCE_YEAR}:{metric}","estimated",{"reference_year":2020,"metric":metric,"estimate":True,"source_page":SOURCE_PAGE}))
        return out


def main():
    parser=argparse.ArgumentParser(); parser.add_argument("--input"); parser.add_argument("--dry-run",action="store_true"); parser.add_argument("--only",action="append",default=[]); args=parser.parse_args()
    url=os.environ.get("SUPABASE_URL"); key=os.environ.get("SUPABASE_SECRET_KEY") or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not args.dry_run and (not url or not key): raise SystemExit("Set SUPABASE_URL and SUPABASE_SECRET_KEY or SUPABASE_SERVICE_ROLE_KEY.")
    warehouse=None if args.dry_run else SupabaseWarehouse(url or "",key or "")
    print(PewReligionImporter(warehouse,input_path=args.input,dry_run=args.dry_run).run(only_keys=set(args.only) or None)); return 0
if __name__=="__main__": raise SystemExit(main())
