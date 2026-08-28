#!/usr/bin/env python3
from __future__ import annotations
import importlib.util, tempfile
from pathlib import Path
from openpyxl import Workbook
from data_pipeline.canonical_countries import CANONICAL_COUNTRY_NAMES


def load_module():
    path=Path(__file__).with_name("import-pew-religion.py")
    spec=importlib.util.spec_from_file_location("pew_importer",path)
    assert spec and spec.loader
    module=importlib.util.module_from_spec(spec); spec.loader.exec_module(module); return module


def make_mixed_book(path:Path, *, include_diversity:bool):
    wb=Workbook(); ws=wb.active
    headers=["Country","Year"]
    labels=["Christian","Muslim","Hindu","Buddhist","Jewish","Other religions","Unaffiliated"]
    for label in labels: headers += [f"{label} population 2020",f"{label} share 2020"]
    if include_diversity: headers += ["Religious diversity index 2020"]
    ws.append(headers)
    for i,(_,name) in enumerate(list(CANONICAL_COUNTRY_NAMES.items())[:120]):
        shares=[40,25,12,8,5,4,6]
        row=[name,2020]
        for j,_ in enumerate(labels): row += [1_000_000+i*1000+j,shares[j]]
        if include_diversity: row += [6.7]
        ws.append(row)
    wb.save(path)


def make_official_layout_book(path:Path):
    """Mirror Pew's real 2025 workbook: generic headers on separate sheets."""
    wb=Workbook()
    counts=wb.active; counts.title="Rounded counts"
    unrounded=wb.create_sheet("Unrounded counts")
    shares=wb.create_sheet("Percentages")
    diversity=wb.create_sheet("Diversity statistics")
    group_headers=["Christians","Muslims","Religiously_unaffiliated","Buddhists","Hindus","Jews","Other_religions"]
    common=["Region","Country","Year","Population"]+group_headers+["Level","Countrycode"]
    counts.append(common); unrounded.append(common); shares.append(common)
    diversity.append(["Region","Country","Year","Diversity rank","RDI score","Diversity level","Buddhists","Christians","Hindus","Jews","Muslims","Other_religions","Religiously_unaffiliated","Level","Country Code"])
    iso_names=list(CANONICAL_COUNTRY_NAMES.items())[:120]
    for i,(iso3,name) in enumerate(iso_names):
        # Deliberately make every count millions and every share <=40 so the
        # semantic guard catches any sheet overwrite immediately.
        population=20_000_000+i*10_000
        group_counts=[4_000_000+i*1000,3_500_000+i*900,2_600_000+i*700,2_800_000+i*600,2_200_000+i*500,1_500_000+i*300,1_400_000+i*200]
        group_shares=[40,25,6,8,12,5,4]
        code=100+i
        counts.append(["Region",name,2020,population]+[round(v,-3) for v in group_counts]+[1,code])
        unrounded.append(["Region",name,2020,population]+[float(v)+0.25 for v in group_counts]+[1,code])
        shares.append(["Region",name,2020,population]+group_shares+[1,code])
        diversity.append(["Region",name,2020,i+1,6.7,"High",8,40,12,5,25,4,6,1,code])
    wb.save(path)


def check(m,path:Path, *, expected_diversity:float, expected_population_floor:float):
    importer=m.PewReligionImporter(None,input_path=str(path),dry_run=True)
    candidates=importer.discover(); assert len(candidates)==15, len(candidates)
    keys={c.rule.key for c in candidates}
    assert {"jewish-share","jewish-population","other-religions-share","other-religions-population","christian-population","religious-diversity"} <= keys
    for c in candidates: assert len(importer.fetch_observations(c))==120, c.rule.key
    christian_share=next(c for c in candidates if c.rule.key=="christian-share")
    christian_total=next(c for c in candidates if c.rule.key=="christian-population")
    assert christian_share.metadata["strategyFamily"]==christian_total.metadata["strategyFamily"]=="religion-christian"
    share_value=importer.fetch_observations(christian_share)[0].value
    total_value=importer.fetch_observations(christian_total)[0].value
    assert 0 <= share_value <= 100, share_value
    assert total_value >= expected_population_floor, total_value
    assert total_value != share_value, (total_value,share_value)
    other_share=next(c for c in candidates if c.rule.key=="other-religions-share")
    other_total=next(c for c in candidates if c.rule.key=="other-religions-population")
    assert "Baha" in other_share.rule.description and "Sikhs" in other_share.rule.description
    assert other_share.metadata.get("groupDefinition")
    assert importer.fetch_observations(other_total)[0].value >= expected_population_floor
    diversity=next(c for c in candidates if c.rule.key=="religious-diversity")
    observed=importer.fetch_observations(diversity)[0].value
    assert abs(observed-expected_diversity)<0.01,(observed,expected_diversity)


def main():
    m=load_module()
    with tempfile.TemporaryDirectory() as d:
        current=Path(d)/"pew-mixed.xlsx"; make_mixed_book(current,include_diversity=True); check(m,current,expected_diversity=6.7,expected_population_floor=1_000_000)
        official=Path(d)/"Religious Composition 2010-2020.xlsx"; make_official_layout_book(official); check(m,official,expected_diversity=6.7,expected_population_floor=1_000_000)
    print("Pew religion 15-category importer fixtures passed, including exact official count/percentage sheet separation and other-religions definition."); return 0


if __name__=="__main__": raise SystemExit(main())
