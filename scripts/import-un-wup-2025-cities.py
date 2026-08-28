#!/usr/bin/env python3
"""GeoStats v16.2.6 UN World Urbanization Prospects 2025 city importer.

Recovers 13 country-level city concepts from the official WUP 2025 city files:
F21 population, F23 population growth, F25 land area, F27 land-area growth,
F29 built-up area, and F32 built-up-area growth.

The 2025 city universe is the official WUP city set with at least 50,000
inhabitants in 2025.  Cross-file derivations require complete stable city-ID
matches for every qualifying city in a country; partial countries are excluded
rather than silently aggregated from an incomplete component set.
"""
from __future__ import annotations

import argparse
import csv
import os
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from data_pipeline.base import WarehouseImporter
from data_pipeline.canonical_countries import CANONICAL_COUNTRY_NAMES, canonical_country_name, country_name_to_iso3
from data_pipeline.models import CandidateDefinition, IndicatorRule, SourceObservation
from data_pipeline.supabase import SupabaseWarehouse

SOURCE_ORG = "United Nations Department of Economic and Social Affairs, Population Division"
SOURCE_DATASET = "World Urbanization Prospects 2025 — Cities"
SOURCE_PAGE = "https://population.un.org/wup/"
DOWNLOAD_PAGE = "https://population.un.org/wup/downloads?tab=Cities"
METHOD = "https://population.un.org/wup/assets/Publications/undesa_pd_2025_wup2025_methodological_report.pdf"
YEAR = 2025
MIN_CITY_THOUSANDS = 50.0

FILES = {
    "population": "WUP2025-F21-DEGURBA-Cities_Pop.xlsx",
    "population_growth": "WUP2025-F23-DEGURBA-Cities_Pop_rate.xlsx",
    "land_area": "WUP2025-F25-DEGURBA-Cities_AREA_km2.xlsx",
    "land_growth": "WUP2025-F27-DEGURBA-Cities_AREA_km2_rate.xlsx",
    "built_area": "WUP2025-F29-DEGURBA-Cities_BU_AREA_km2.xlsx",
    "built_growth": "WUP2025-F32-DEGURBA-Cities_BU_km2_rate.xlsx",
}
BASE_DOWNLOAD = "https://population.un.org/wup/assets/Download/Cities/"
FILE_CODE = {
    "population": "F21", "population_growth": "F23", "land_area": "F25",
    "land_growth": "F27", "built_area": "F29", "built_growth": "F32",
}


def norm(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def parse_float(value: object) -> float | None:
    if value in (None, "", "..", "…", "—", "-"):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def _canonical_iso3(raw_iso3: object, raw_country: object) -> str | None:
    iso3 = str(raw_iso3 or "").strip().upper()
    if iso3 not in CANONICAL_COUNTRY_NAMES:
        iso3 = country_name_to_iso3(str(raw_country or "")) or ""
    return iso3 if iso3 in CANONICAL_COUNTRY_NAMES else None


def _header_map(row: list[object]) -> dict[str, int]:
    names = [norm(v) for v in row]
    out: dict[str, int] = {}
    for i, name in enumerate(names):
        if name in {"iso3", "iso3code", "iso3alphacode", "alpha3code"}:
            out.setdefault("iso3", i)
        if any(token in name for token in ("countryorarea", "countryarea", "location")):
            out.setdefault("country", i)
        if name in {"cityid", "citycode", "urbancentreid", "urbancenterid", "cityidentifier"}:
            out.setdefault("city_id", i)
        if name in {"city", "cityname", "urbanagglomeration", "urbanagglomerationname", "urbancentre", "urbancenter"}:
            out.setdefault("city", i)
        if name in {"year", "time", "timeperiod"}:
            out.setdefault("year", i)
        if name in {"value", "estimate", "population", "area", "rate"}:
            out.setdefault("value", i)
    return out


def _parse_city_xlsx(path: Path) -> list[tuple[str, str, str, str, int, float]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        for header_idx, header in enumerate(rows[:80]):
            mapping = _header_map(header)
            year_columns = {
                i: int(str(value).strip())
                for i, value in enumerate(header)
                if str(value or "").strip().isdigit() and 1975 <= int(str(value).strip()) <= 2050
            }
            if not {"iso3", "city_id"}.issubset(mapping) or ("year" not in mapping and not year_columns):
                continue
            if "city" not in mapping:
                raise RuntimeError(f"WUP city workbook {path.name} has a city ID but no city-name field.")
            long_layout = "year" in mapping
            if long_layout and "value" not in mapping:
                names = [norm(v) for v in header]
                candidate_cols = [i for i, name in enumerate(names) if any(t in name for t in ("population", "area", "rate", "value", "estimate"))]
                candidate_cols = [i for i in candidate_cols if i not in set(mapping.values())]
                if len(candidate_cols) == 1:
                    mapping["value"] = candidate_cols[0]
            if long_layout and "value" not in mapping:
                raise RuntimeError(f"WUP city workbook {path.name} has long-format years but no unique value column.")
            parsed = []
            for row in rows[header_idx + 1:]:
                def cell(key: str) -> object:
                    i = mapping.get(key)
                    return row[i] if i is not None and i < len(row) else None
                iso3 = _canonical_iso3(cell("iso3"), cell("country"))
                if not iso3:
                    continue
                city_id = str(cell("city_id") or "").strip()
                city = str(cell("city") or "").strip()
                if not city_id or not city:
                    continue
                country = canonical_country_name(iso3, str(cell("country") or iso3))
                if long_layout:
                    try:
                        year = int(float(str(cell("year"))))
                    except Exception:
                        continue
                    value = parse_float(cell("value"))
                    if value is not None:
                        parsed.append((iso3, country, city_id, city, year, value))
                else:
                    for col, year in year_columns.items():
                        if col >= len(row):
                            continue
                        value = parse_float(row[col])
                        if value is not None:
                            parsed.append((iso3, country, city_id, city, year, value))
            if parsed:
                return parsed
    raise RuntimeError(f"Could not find an unambiguous WUP city data sheet in {path.name}.")


def _load_normalized_csv(path: Path):
    metric_columns = {
        "population": "populationthousands",
        "population_growth": "populationgrowthpct",
        "land_area": "landareakm2",
        "land_growth": "landgrowthpct",
        "built_area": "builtareakm2",
        "built_growth": "builtgrowthpct",
    }
    data = {metric: {} for metric in metric_columns}
    names: dict[tuple[str, str], tuple[str, str]] = {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fields = {norm(v): v for v in (reader.fieldnames or [])}
        for required in ("iso3", "cityid", "city", "year"):
            if required not in fields:
                raise RuntimeError(f"Normalized WUP city CSV is missing {required}.")
        for col in metric_columns.values():
            if col not in fields:
                raise RuntimeError(f"Normalized WUP city CSV is missing {col}.")
        for row in reader:
            mapped = {norm(k): v for k, v in row.items()}
            iso3 = _canonical_iso3(mapped.get("iso3"), mapped.get("country"))
            city_id = str(mapped.get("cityid") or "").strip()
            city = str(mapped.get("city") or "").strip()
            try:
                year = int(float(str(mapped.get("year"))))
            except Exception:
                continue
            if not iso3 or not city_id or not city:
                continue
            key = (iso3, city_id, year)
            name_key = (iso3, city_id)
            prior_name = names.get(name_key)
            current_name = (canonical_country_name(iso3, str(mapped.get("country") or iso3)), city)
            if prior_name and prior_name != current_name:
                raise RuntimeError(f"Conflicting WUP city identity for {name_key}: {prior_name} vs {current_name}")
            names[name_key] = current_name
            for metric, col in metric_columns.items():
                value = parse_float(mapped.get(col))
                if value is None:
                    continue
                if key in data[metric] and abs(data[metric][key] - value) > 1e-12:
                    raise RuntimeError(f"Conflicting duplicate WUP city observation for {metric} {key}")
                data[metric][key] = value
    return data, names


def load_input(input_path: str):
    path = Path(input_path)
    if path.is_file() and path.suffix.lower() == ".csv":
        return _load_normalized_csv(path)
    if not path.is_dir():
        raise RuntimeError("--input must be a normalized CSV or directory containing official WUP 2025 city workbooks.")
    data = {metric: {} for metric in FILES}
    names: dict[tuple[str, str], tuple[str, str]] = {}
    for metric, filename in FILES.items():
        file_path = path / filename
        if not file_path.exists():
            raise RuntimeError(f"Missing official WUP city file: {filename}")
        for iso3, country, city_id, city, year, value in _parse_city_xlsx(file_path):
            key = (iso3, city_id, year)
            name_key = (iso3, city_id)
            prior = names.get(name_key)
            if prior and prior != (country, city):
                raise RuntimeError(f"Conflicting WUP city identity across official files for {name_key}")
            names[name_key] = (country, city)
            if key in data[metric] and abs(data[metric][key] - value) > 1e-12:
                raise RuntimeError(f"Conflicting duplicate WUP city observation for {metric} {key}")
            data[metric][key] = value
    return data, names


SPECS = {
    "city-count": ("Most cities", "Number of WUP cities with at least 50,000 inhabitants in 2025.", "cities", "total", "high"),
    "million-city-count": ("Most million-person cities", "Number of WUP cities with at least 1 million inhabitants in 2025.", "cities", "total", "high"),
    "megacity-count": ("Most megacities", "Number of WUP cities with at least 10 million inhabitants in 2025.", "cities", "total", "high"),
    "largest-city": ("Largest city population", "Population of the country's largest WUP city in 2025.", "people", "total", "high"),
    "largest-city-concentration": ("Highest largest-city concentration", "Largest single city's share of the country's population living in WUP cities.", "% of city population", "percentage", "high"),
    "city-density": ("Highest overall city density", "Population of all WUP cities divided by their combined city land area in 2025.", "people/km²", "rate", "high"),
    "total-city-area": ("Largest total city land area", "Combined land area of WUP cities with at least 50,000 inhabitants in 2025.", "km²", "total", "high"),
    "total-built-area": ("Largest built-up city area", "Combined built-up area inside WUP cities with at least 50,000 inhabitants in 2025.", "km²", "total", "high"),
    "built-area-per-person": ("Most built-up city area per resident", "Combined built-up city area per resident of WUP cities in 2025.", "m²/person", "per_capita", "high"),
    "built-share": ("Highest built-up share of city land", "Built-up area as a share of the combined land area of WUP cities in 2025.", "% of city land", "percentage", "high"),
    "population-growth": ("Fastest average city population growth", "Population-weighted average annual growth rate across WUP cities in 2025.", "% per year", "rate", "high"),
    "land-growth": ("Fastest city land expansion", "Land-area-weighted average annual growth rate across WUP cities in 2025.", "% per year", "rate", "high"),
    "built-growth": ("Fastest built-up city expansion", "Built-up-area-weighted average annual growth rate across WUP cities in 2025.", "% per year", "rate", "high"),
}


class Importer(WarehouseImporter):
    source_organization = SOURCE_ORG
    source_dataset = SOURCE_DATASET
    source_slug = "unwupcities2025"

    def __init__(self, warehouse, input_path: str, dry_run: bool = False):
        super().__init__(warehouse, dry_run=dry_run)
        self.data, self.names = load_input(input_path)
        self._cache: dict[str, dict[str, float]] = {}

    def _active_cities(self, iso3: str) -> list[str]:
        return sorted({city_id for (country, city_id, year), value in self.data["population"].items() if country == iso3 and year == YEAR and value >= MIN_CITY_THOUSANDS})

    def _country_value(self, key: str, iso3: str) -> float | None:
        cities = self._active_cities(iso3)
        if not cities:
            return None
        def series(metric: str) -> list[float] | None:
            vals = []
            for city_id in cities:
                value = self.data[metric].get((iso3, city_id, YEAR))
                if value is None:
                    return None
                vals.append(value)
            return vals
        pops = series("population")
        if pops is None:
            return None
        if key == "city-count": return float(len(cities))
        if key == "million-city-count": return float(sum(v >= 1000 for v in pops))
        if key == "megacity-count": return float(sum(v >= 10000 for v in pops))
        total_pop_thousands = sum(pops)
        if total_pop_thousands <= 0:
            return None
        if key == "largest-city": return max(pops) * 1000.0
        if key == "largest-city-concentration": return 100.0 * max(pops) / total_pop_thousands
        land = series("land_area")
        if key in {"city-density", "total-city-area", "built-share", "land-growth"} and land is None:
            return None
        built = series("built_area")
        if key in {"total-built-area", "built-area-per-person", "built-share", "built-growth"} and built is None:
            return None
        if land is not None and any(v <= 0 for v in land):
            return None
        if built is not None and any(v < 0 for v in built):
            return None
        total_land = sum(land or [])
        total_built = sum(built or [])
        if key == "city-density": return total_pop_thousands * 1000.0 / total_land if total_land > 0 else None
        if key == "total-city-area": return total_land
        if key == "total-built-area": return total_built
        if key == "built-area-per-person": return total_built * 1_000_000.0 / (total_pop_thousands * 1000.0)
        if key == "built-share":
            if total_land <= 0 or total_built > total_land * 1.000001:
                return None
            return 100.0 * total_built / total_land
        if key == "population-growth":
            growth = series("population_growth")
            if growth is None: return None
            return sum(w * g for w, g in zip(pops, growth)) / total_pop_thousands
        if key == "land-growth":
            growth = series("land_growth")
            if growth is None or total_land <= 0: return None
            return sum(w * g for w, g in zip(land or [], growth)) / total_land
        if key == "built-growth":
            growth = series("built_growth")
            if growth is None or total_built <= 0: return None
            return sum(w * g for w, g in zip(built or [], growth)) / total_built
        raise KeyError(key)

    def _values_for(self, key: str) -> dict[str, float]:
        if key not in self._cache:
            vals = {}
            for iso3 in CANONICAL_COUNTRY_NAMES:
                value = self._country_value(key, iso3)
                if value is not None:
                    vals[iso3] = value
            self._cache[key] = vals
        return self._cache[key]

    def discover(self) -> list[CandidateDefinition]:
        out = []
        for key, (title, desc, unit, value_type, direction) in SPECS.items():
            eligible = sorted(self._values_for(key))
            if len(eligible) < 12:
                raise RuntimeError(f"WUP city concept {key} has only {len(eligible)} eligible countries in the supplied snapshot.")
            component_tables = {
                "city-count": ["F21"], "million-city-count": ["F21"], "megacity-count": ["F21"],
                "largest-city": ["F21"], "largest-city-concentration": ["F21"],
                "city-density": ["F21","F25"], "total-city-area": ["F25"],
                "total-built-area": ["F29"], "built-area-per-person": ["F21","F29"],
                "built-share": ["F25","F29"], "population-growth": ["F21","F23"],
                "land-growth": ["F25","F27"], "built-growth": ["F29","F32"],
            }[key]
            rule = IndicatorRule(
                key=key,title=title,description=desc,plain_language_description=desc,
                technical_definition=desc,unit_explanation=unit,family="Cities",icon="🏙️",
                unit=unit,value_type=value_type,ranking_direction=direction,include=(key,),
                min_coverage=min(100, len(eligible)), evidence_tier="A",source_priority=5,
                specificity_score=98,recognizability_score=95,understandability_score=94,fun_score=88,
            )
            out.append(CandidateDefinition(rule, f"WUP2025:{'+'.join(component_tables)}:{key}:{YEAR}", title, SOURCE_PAGE, {
                "source_page_url": SOURCE_PAGE,
                "methodology_url": METHOD,
                "dataset_release": "World Urbanization Prospects 2025 Revision",
                "source_query": {"tables": component_tables, "year": YEAR, "city_threshold_people": 50000},
                "input_datasets": [FILES[{"F21":"population","F23":"population_growth","F25":"land_area","F27":"land_growth","F29":"built_area","F32":"built_growth"}[code]] for code in component_tables],
                "measurementType": "percentage" if value_type == "percentage" else ("per_capita" if value_type == "per_capita" else ("total" if value_type == "total" else "rate")),
                "broadDomain": "human-geography", "knowledgeCluster": "cities", "strategyFamily": "cities-2025",
                "manual_review_required": True,
                "eligible_universe_type": "defined_subset",
                "eligible_universe_rule": "GeoStats canonical countries with at least one WUP 2025 city >=50,000 population and complete official component data required for this concept.",
                "eligible_country_count": len(eligible),
                "eligible_country_iso3": eligible,
                "eligible_universe_selector": f"WUP 2025 F21 threshold plus complete components for {key}",
                "excluded_country_reason": "No qualifying WUP 2025 city or incomplete required city-level component data.",
                "comparison_year_policy": "Use 2025 city observations only; do not mix projection years.",
                "derivation_method": desc,
                "v16_2_6_content_reviewed": True,
            }))
        return out

    def fetch_observations(self, candidate: CandidateDefinition) -> list[SourceObservation]:
        values = self._values_for(candidate.rule.key)
        out = []
        for iso3 in candidate.metadata["eligible_country_iso3"]:
            value = values.get(iso3)
            if value is None:
                raise RuntimeError(f"Eligible WUP city country {iso3} lost its derived value for {candidate.rule.key}.")
            if candidate.rule.value_type == "percentage" and not (0 <= value <= 100):
                raise RuntimeError(f"WUP city percentage outside 0-100 for {candidate.rule.key} {iso3}: {value}")
            out.append(SourceObservation(iso3, canonical_country_name(iso3, iso3), YEAR, value, SOURCE_PAGE,
                f"{candidate.source_indicator_code}:{iso3}", "official", {
                    "reference_year": YEAR,
                    "qualifying_city_count": len(self._active_cities(iso3)),
                    "city_threshold_people": 50000,
                }))
        return out

    def category_id(self, candidate: CandidateDefinition) -> str:
        return f"unwupcities2025:{candidate.rule.key}"


def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--input', required=True); ap.add_argument('--dry-run', action='store_true'); args=ap.parse_args()
    url=os.getenv('SUPABASE_URL'); key=os.getenv('SUPABASE_SECRET_KEY') or os.getenv('SUPABASE_SERVICE_ROLE_KEY')
    if not args.dry_run and (not url or not key): raise SystemExit('Set Supabase secrets.')
    print(Importer(None if args.dry_run else SupabaseWarehouse(url,key), args.input, args.dry_run).run())

if __name__=='__main__': main()
